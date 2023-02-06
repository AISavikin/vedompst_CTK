import customtkinter as ctk
from tkinter import messagebox, ttk, END
from openpyxl import load_workbook, styles
from utils import *
import os


class SheetWindow(Mixin, ctk.CTkToplevel):
    def __init__(self, parent, group, month_name: str):
        super().__init__(parent)
        self.title('Ведомости')
        self.group = group
        self.month_name = month_name
        self.font = parent.font
        w = 610
        self.set_geometry(w)

        self.table_frame = TableFrame(self)
        self.control_frame = ControlFrame(self)

        self.control_frame.pack()
        self.table_frame.pack()

    def close_sheet(self):
        month_num = self.get_month_num(self.month_name)
        absent_dicts = self.table_frame.absent_dicts
        work_days = self.table_frame.work_days
        path_sheet = self.get_settings()['PATH_SHEET']
        path_screenshot = self.get_settings()['PATH_SCREENSHOT']
        group = self.control_frame.label.cget('text')
        weekend = self.get_weekend(month_num)
        close_day = self.control_frame.close_day.get()
        month_name = self.month_name
        year = self.get_year(month_num)
        CloseSheet(absent_dicts, work_days, path_sheet, path_screenshot, group, weekend, month_name, year,
                   close_day).save()
        messagebox.showinfo('Сохранено', message=f'Ведомость {group} сохранена, успешно.')
        self.destroy()


class ControlFrame(ctk.CTkFrame):
    def __init__(self, master: SheetWindow):
        super().__init__(master)
        group_nums = [f'Группа {num}' for num in set(i.group for i in Student.select()) if str(num) != master.group]
        self.label = ctk.CTkLabel(self, text=f'Группа {master.group} {master.month_name}', font=master.font)
        self.label.grid(row=0, columnspan=2, padx=5, pady=7)
        self.ext_group = ctk.CTkComboBox(self, values=group_nums, font=master.font)
        self.ext_group.grid(row=1, column=0, padx=5, pady=7, sticky='nesw')
        ctk.CTkButton(self, text='Добавить', command=master.table_frame.add_group,
                      font=master.font).grid(row=1, column=1, padx=5, pady=7, sticky='nesw')
        self.close_day = ctk.CTkComboBox(self, values=[str(i) for i in range(1, 31)], font=master.font, width=80)
        self.close_day.grid(row=2, column=0, padx=5, pady=7)
        self.close_day.set(f'{master.date:%d}')
        ctk.CTkButton(self, text='Закрыть ведомость', command=master.close_sheet,
                      font=master.font).grid(row=2, column=1, sticky='nesw', padx=5, pady=7)


class TableFrame(ctk.CTkFrame):
    def __init__(self, master: SheetWindow):
        super().__init__(master)
        self.master: SheetWindow = master
        self.font = master.font
        self.group = master.group
        self.month_name = master.month_name
        self.month_num = master.get_month_num(self.month_name)
        self.year = master.get_year(self.month_num)
        self.work_days = master.get_work_days(self.month_name)

        self.absent_dicts = [self.get_absent_dict(self.group)]
        self.gen_table(self.absent_dicts)

    def get_absent_dict(self, group: Union[str, int]) -> dict:
        kids = Student.select().filter(group=group).order_by(Student.name)
        work_days_from_db = [i.day for i in Attendance.select().filter(month=self.month_num, student_id=kids[0],
                                                                       year=self.year)]
        self.work_days = list(set(work_days_from_db) | set(self.work_days))
        absent_dict = {}
        for kid in kids:
            absent_dict[kid.name] = {day: '' for day in self.work_days}
            row = [(i.day, i.absent) for i in
                   Attendance.select().filter(month=self.month_num, student_id=kid.id, year=self.year)]
            for day in row:
                absent_dict[kid.name][day[0]] = day[1]

        return absent_dict

    def gen_table(self, absent_dicts):
        ttk.Style().configure('Treeview', rowheight=25)
        columns = ('№', 'name') + tuple(str(i) for i in self.work_days) + ('н', 'б')
        self.table = ttk.Treeview(self, columns=columns, show='headings')
        for i in columns:
            self.table.heading(i, text=i)
            self.table.column(i, width=30, anchor='n')
        self.table.column('name', width=220, anchor='n')
        self.table.heading('name', text='Фамилия, Имя')
        self.paste(absent_dicts)
        self.table.pack(padx=10, pady=10)

    def paste(self, absent_dicts):
        self.table.tag_configure('white', foreground='white', background='#1a1a1a', font=self.font)
        self.table.configure(height=sum(len(d) for d in absent_dicts))
        num = 1
        for absent_dict in absent_dicts:
            for kid in sorted(absent_dict.keys()):
                title = (str(num), kid)
                absents = tuple(absent_dict[kid].get(i, '') for i in self.work_days)
                count = (absents.count('н'), absents.count('б'))
                values = title + absents + count
                self.table.insert("", END, values=values, tags='white')
                num += 1

    def add_group(self):
        groups = self.master.control_frame.label.cget('text').split()[1:-1]
        ext_group = self.master.control_frame.ext_group.get().split()[1]
        groups.append(ext_group)
        ext_absent_dict = self.get_absent_dict(ext_group)
        self.absent_dicts.append(ext_absent_dict)
        self.master.control_frame.label.configure(text=f'Группа {",".join(groups)} {self.month_name}')
        self.table.destroy()
        self.gen_table(self.absent_dicts)


class CloseSheet:

    def __init__(self, absent_dicts: list[dict],
                 work_days: list[str],
                 path_sheet: str,
                 path_screenshot: str,
                 group: str,
                 weekends: list[int],
                 month_name: str,
                 year: int,
                 close_day: str):
        self.absent_dicts = absent_dicts
        self.work_days = work_days
        self.path_sheet = path_sheet
        self.path_screenshot = path_screenshot
        self.weekends = weekends
        self.month_name = month_name
        self.year = year
        self.group = group
        self.close_day = close_day
        self.filename = f'{path_sheet}{group}.xlsx'

    def check_directory(self):
        """Проверяет существует ли директория, и создает если необходимо"""
        if not os.path.exists(self.path_sheet):
            os.makedirs(self.path_sheet)

    def write_data(self):
        wb = load_workbook('Template.xlsx')
        ws = wb['Посещаемость']

        cols = [2] + [int(i) + 4 for i in self.work_days] + [36, 38]
        row_num = 16
        for absent_dict in self.absent_dicts:
            for kid in sorted(absent_dict.keys()):
                row_data = [kid] + [absent_dict[kid].get(i, '') for i in self.work_days]
                row_data += [row_data.count('н'), row_data.count('б')]
                for indx, column in enumerate(cols):
                    ws.cell(row=row_num, column=column).value = row_data[indx]
                row_num += 1
        wb.save(self.filename)

    def save(self):
        self.check_directory()
        self.write_data()
        work_book = load_workbook(self.filename)
        work_sheet = work_book['Посещаемость']
        work_days_num = work_sheet['aj16'].value + work_sheet['al16'].value
        self.write_service_information(work_sheet, work_days_num)
        self.colorize_weekend(work_sheet)
        work_book.save(self.filename)

    def write_service_information(self, ws, work_days_num):

        ws['N3'].value = self.month_name
        ws['AA43'].value = self.month_name
        ws['C5'].value = ' '.join(self.group.split()[:-1])
        ws['V3'].value = self.year
        ws['AG43'].value = self.year
        ws['Y43'].value = self.close_day
        ws['C7'].value = work_days_num

    def colorize_weekend(self, ws):
        for row in range(16, 39):
            for col in self.weekends:
                ws.cell(row=row, column=col + 4).fill = styles.PatternFill(start_color='5E5E5E', fill_type='solid')
